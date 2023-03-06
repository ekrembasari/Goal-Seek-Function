import openpyxl
from openpyxl import Workbook
from decimal import Decimal

# Import the openpyxl library and the Workbook class from the openpyxl module.
# Import the Decimal class from the decimal module.

def goal_seek(f, a, b, tol=1e-8):
    """
    Finds a root of the function f in the interval [a, b] to within a tolerance tol.

    f: a function that takes a single argument
    a: left endpoint of the interval
    b: right endpoint of the interval
    tol: tolerance (default: 1e-8)
    """
    
    while abs(f(a)) > tol:
        x = float(Decimal((a + b)) / Decimal(2))
        if f(x) == 0:
            return x
        elif Decimal(f(a)) * Decimal(f(x)) < 0:
            b = x
        else:
            a = x
    return (a + b) / 2

# Define the goal_seek() function. This function uses the bisection method to find a root of the given function f.

def f(x):
    r1 = 2
    r2 = 3

    # Calculate the values of F column and returns the value of the function in F102.
    # r1 and r2 are the row numbers of the cells that are being read or written to.

    for i in range(100):
        ws.cell(r2, 7).value = Decimal(ws.cell(r1, 7).value) * Decimal((Decimal(1) + Decimal(ws.cell(r2, 5).value))) + (Decimal(ws.cell(r2, 2).value) * Decimal(x) - Decimal(ws.cell(r2, 3).value))
        r1 += 1
        r2 += 1
    return ws['G102'].value

# Load the Excel workbook
wb = openpyxl.load_workbook('test.xlsx')

# Select the worksheet that contains the goal seek function
ws = wb['Sheet1']

a = float(ws['I6'].value)
b = float(a + 1)

# Calculate the desired value of I6
result = goal_seek(f, a, b, tol=1e-8)
print("I6 must be equal to: " + str(result))

# Print the rounded value of G102
print(round(ws['G102'].value))
